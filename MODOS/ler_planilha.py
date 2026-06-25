import csv
import json
from pathlib import Path


def _normalizar_lista(valor):
    if valor is None:
        return []
    if isinstance(valor, str):
        texto = valor.strip()
        if not texto:
            return []
        try:
            carregado = json.loads(texto)
            if isinstance(carregado, list):
                return carregado
        except Exception:
            pass
        return [parte.strip() for parte in texto.split(",") if parte.strip()]
    if isinstance(valor, list):
        return valor
    return [valor]


def _nome_arquivo_seguro(nome_arquivo: str, extensao_padrao: str) -> str | None:
    nome = (nome_arquivo or "").strip()
    if (
        not nome
        or Path(nome).is_absolute()
        or any(parte == ".." for parte in Path(nome).parts)
        or "/" in nome
        or "\\" in nome
    ):
        return None
    if not Path(nome).suffix:
        nome += extensao_padrao
    return nome


def _linhas_csv(caminho: Path) -> tuple[list[str], list[dict]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = caminho.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        texto = caminho.read_text(encoding="utf-8", errors="replace")

    amostra = texto[:4096]
    try:
        dialecto = csv.Sniffer().sniff(amostra)
    except Exception:
        dialecto = csv.excel

    reader = csv.DictReader(texto.splitlines(), dialect=dialecto)
    colunas = list(reader.fieldnames or [])
    linhas = [dict(row) for row in reader]
    return colunas, linhas


def _linhas_xlsx(caminho: Path, aba: str = "") -> tuple[list[str], list[dict]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("Para ler .xlsx, instale a dependencia openpyxl.") from exc

    workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    worksheet = workbook[aba] if aba and aba in workbook.sheetnames else workbook.active
    linhas_iter = worksheet.iter_rows(values_only=True)
    cabecalho = next(linhas_iter, None)
    if not cabecalho:
        return [], []

    colunas = [str(valor).strip() if valor is not None else f"coluna_{i + 1}" for i, valor in enumerate(cabecalho)]
    linhas = []
    for valores in linhas_iter:
        linha = {}
        for i, coluna in enumerate(colunas):
            valor = valores[i] if i < len(valores) else ""
            linha[coluna] = "" if valor is None else valor
        linhas.append(linha)
    workbook.close()
    return colunas, linhas


def ler_planilha(caminho: str, filtro_coluna: str = "", filtro_valor: str = "", limite_linhas: int = 20, aba: str = "") -> str:
    try:
        path = Path(caminho).expanduser().resolve()
        if not path.is_file():
            return f"Arquivo nao encontrado: {path}"

        ext = path.suffix.lower()
        if ext == ".csv":
            colunas, linhas = _linhas_csv(path)
        elif ext == ".xlsx":
            colunas, linhas = _linhas_xlsx(path, aba=aba)
        else:
            return "Formato nao suportado. Use .csv ou .xlsx."

        total_original = len(linhas)
        if filtro_coluna and filtro_valor:
            if filtro_coluna not in colunas:
                return f"Coluna '{filtro_coluna}' nao encontrada. Colunas: {', '.join(colunas)}"
            alvo = str(filtro_valor).casefold()
            linhas = [linha for linha in linhas if alvo in str(linha.get(filtro_coluna, "")).casefold()]

        limite = max(1, min(int(limite_linhas or 20), 100))
        preview = linhas[:limite]
        partes = [
            f"Planilha: {path.name}",
            f"Colunas ({len(colunas)}): {', '.join(colunas) if colunas else '(sem cabecalho)'}",
            f"Linhas: {len(linhas)}" + (f" de {total_original} apos filtro" if len(linhas) != total_original else ""),
            "",
            "Previa:",
        ]

        if not preview:
            partes.append("(nenhuma linha para mostrar)")
        else:
            for idx, linha in enumerate(preview, 1):
                valores = [f"{coluna}={linha.get(coluna, '')}" for coluna in colunas[:12]]
                partes.append(f"{idx}. " + " | ".join(valores))

        if len(linhas) > limite:
            partes.append(f"... mais {len(linhas) - limite} linhas ocultas pelo limite.")
        return "\n".join(partes)
    except Exception as e:
        return f"Erro ao ler planilha: {e}"


def criar_planilha(nome_arquivo: str, colunas, linhas=None, formato: str = "xlsx") -> str:
    try:
        formato = (formato or "xlsx").strip().lower()
        if formato not in ("xlsx", "csv"):
            return "Formato invalido. Use 'xlsx' ou 'csv'."

        nome_seguro = _nome_arquivo_seguro(nome_arquivo, f".{formato}")
        if not nome_seguro:
            return "Use apenas o nome do arquivo, sem caminho, subpastas ou '..'."

        colunas = [str(coluna) for coluna in _normalizar_lista(colunas)]
        if not colunas:
            return "Informe pelo menos uma coluna."

        linhas = _normalizar_lista(linhas)
        pasta_saida = Path(__file__).resolve().parent.parent / "DocumentosCriados"
        pasta_saida.mkdir(parents=True, exist_ok=True)
        caminho_final = pasta_saida / nome_seguro

        if formato == "csv" or caminho_final.suffix.lower() == ".csv":
            if caminho_final.suffix.lower() != ".csv":
                caminho_final = caminho_final.with_suffix(".csv")
            with caminho_final.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=colunas)
                writer.writeheader()
                for linha in linhas:
                    if isinstance(linha, dict):
                        writer.writerow({coluna: linha.get(coluna, "") for coluna in colunas})
                    else:
                        valores = _normalizar_lista(linha)
                        writer.writerow({coluna: valores[i] if i < len(valores) else "" for i, coluna in enumerate(colunas)})
        else:
            try:
                import openpyxl
            except ImportError:
                return "Para criar .xlsx, instale a dependencia openpyxl."

            if caminho_final.suffix.lower() != ".xlsx":
                caminho_final = caminho_final.with_suffix(".xlsx")
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Planilha"
            worksheet.append(colunas)
            for linha in linhas:
                if isinstance(linha, dict):
                    worksheet.append([linha.get(coluna, "") for coluna in colunas])
                else:
                    valores = _normalizar_lista(linha)
                    worksheet.append([valores[i] if i < len(valores) else "" for i, _ in enumerate(colunas)])
            workbook.save(caminho_final)

        return f"Planilha criada em: {caminho_final}"
    except Exception as e:
        return f"Erro ao criar planilha: {e}"


def register(tool_map, function_declarations):
    tool_map["ler_planilha"] = ler_planilha
    tool_map["criar_planilha"] = criar_planilha
    function_declarations.append({
        "name": "ler_planilha",
        "description": "Le uma planilha .csv ou .xlsx, resume colunas e mostra uma previa, com filtro opcional por coluna/valor.",
        "parameters": {
            "type": "object",
            "properties": {
                "caminho": {"type": "string", "description": "Caminho do arquivo .csv ou .xlsx."},
                "filtro_coluna": {"type": "string", "description": "Coluna opcional para filtrar."},
                "filtro_valor": {"type": "string", "description": "Valor opcional a procurar na coluna filtrada."},
                "limite_linhas": {"type": "integer", "description": "Maximo de linhas na previa, ate 100."},
                "aba": {"type": "string", "description": "Nome da aba do .xlsx. Se vazio usa a aba ativa."}
            },
            "required": ["caminho"]
        }
    })
    function_declarations.append({
        "name": "criar_planilha",
        "description": "Cria uma planilha .xlsx ou .csv em DocumentosCriados a partir de colunas e linhas.",
        "parameters": {
            "type": "object",
            "properties": {
                "nome_arquivo": {"type": "string", "description": "Nome do arquivo, sem caminho."},
                "colunas": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Lista de nomes das colunas."
                },
                "linhas": {
                    "type": "array",
                    "items": {"type": "object", "additionalProperties": {"type": "string"}},
                    "description": "Linhas como lista de objetos {coluna: valor}. Tambem aceita listas simples."
                },
                "formato": {"type": "string", "enum": ["xlsx", "csv"], "description": "Formato de saida."}
            },
            "required": ["nome_arquivo", "colunas"]
        }
    })
